from __future__ import annotations

import os
from datetime import datetime
from typing import Any

import pandas as pd

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ==========================================================
# Corporate Theme
# ==========================================================

PRIMARY_BLUE = "2563EB"
DARK_BLUE = "1E3A8A"
DARK_TEXT = "0F172A"
SECONDARY_TEXT = "475569"
MUTED_TEXT = "64748B"

WHITE = "FFFFFF"
LIGHT_BLUE = "EFF6FF"
LIGHT_GRAY = "F8FAFC"
MEDIUM_GRAY = "E2E8F0"
BORDER_GRAY = "CBD5E1"

SUCCESS_GREEN = "15803D"
LIGHT_GREEN = "DCFCE7"

WARNING_ORANGE = "C2410C"
LIGHT_ORANGE = "FFEDD5"

DANGER_RED = "B91C1C"
LIGHT_RED = "FEE2E2"


# ==========================================================
# Reusable Styles
# ==========================================================

THIN_BORDER_SIDE = Side(
    style="thin",
    color=BORDER_GRAY
)

THIN_BORDER = Border(
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE
)

TITLE_FILL = PatternFill(
    fill_type="solid",
    fgColor=DARK_BLUE
)

SECTION_FILL = PatternFill(
    fill_type="solid",
    fgColor=PRIMARY_BLUE
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor=PRIMARY_BLUE
)

LIGHT_HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor=LIGHT_BLUE
)

ALTERNATE_ROW_FILL = PatternFill(
    fill_type="solid",
    fgColor=LIGHT_GRAY
)

SUCCESS_FILL = PatternFill(
    fill_type="solid",
    fgColor=LIGHT_GREEN
)

WARNING_FILL = PatternFill(
    fill_type="solid",
    fgColor=LIGHT_ORANGE
)

DANGER_FILL = PatternFill(
    fill_type="solid",
    fgColor=LIGHT_RED
)


# ==========================================================
# Workbook Setup
# ==========================================================

def create_professional_workbook() -> Workbook:
    """
    Create a new Excel workbook and remove the default sheet.

    Worksheets will be added individually by the report-building
    functions.
    """

    workbook = Workbook()

    default_sheet = workbook.active

    if default_sheet is not None:
        workbook.remove(
            default_sheet
        )

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    return workbook


# ==========================================================
# General Utilities
# ==========================================================

def ensure_output_directory(
    output_path: str
) -> None:
    """
    Create the output directory when it does not exist.
    """

    directory = os.path.dirname(
        output_path
    )

    if directory:
        os.makedirs(
            directory,
            exist_ok=True
        )


def safe_sheet_name(
    name: str
) -> str:
    """
    Return a valid Excel worksheet name.

    Excel sheet names:
    - cannot exceed 31 characters;
    - cannot contain \\ / * ? : [ ].
    """

    invalid_characters = (
        "\\",
        "/",
        "*",
        "?",
        ":",
        "[",
        "]"
    )

    safe_name = str(
        name or "Report"
    )

    for character in invalid_characters:
        safe_name = safe_name.replace(
            character,
            "-"
        )

    safe_name = safe_name.strip()

    if not safe_name:
        safe_name = "Report"

    return safe_name[:31]


def safe_excel_value(
    value: Any
) -> Any:
    """
    Convert pandas, NumPy and unsupported values into types that
    openpyxl can safely write to Excel.
    """

    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except (
        TypeError,
        ValueError
    ):
        pass

    if isinstance(
        value,
        pd.Timestamp
    ):
        return value.to_pydatetime()

    if hasattr(
        value,
        "item"
    ):
        try:
            return value.item()
        except (
            ValueError,
            TypeError,
            AttributeError
        ):
            pass

    if isinstance(
        value,
        (
            str,
            int,
            float,
            bool,
            datetime
        )
    ):
        return value

    return str(
        value
    )


def format_integer(
    value: Any
) -> str:
    """
    Format a value as an integer with thousands separators.
    """

    try:
        return f"{int(value):,}"
    except (
        TypeError,
        ValueError
    ):
        return str(
            value
        )


def format_percentage(
    value: Any,
    decimal_places: int = 2
) -> str:
    """
    Format a numeric value as a percentage string.
    """

    try:
        return (
            f"{float(value):.{decimal_places}f}%"
        )
    except (
        TypeError,
        ValueError
    ):
        return str(
            value
        )


# ==========================================================
# Dataset Metrics
# ==========================================================

def calculate_dataset_metrics(
    df: pd.DataFrame
) -> dict[str, Any]:
    """
    Calculate the core metrics shown in the Executive Summary.
    """

    total_rows = int(
        len(df)
    )

    total_columns = int(
        len(df.columns)
    )

    missing_values = int(
        df.isna()
        .sum()
        .sum()
    )

    duplicate_rows = int(
        df.duplicated()
        .sum()
    )

    numeric_columns = int(
        len(
            df.select_dtypes(
                include="number"
            ).columns
        )
    )

    categorical_columns = (
        total_columns
        - numeric_columns
    )

    total_cells = max(
        total_rows
        * total_columns,
        1
    )

    missing_percentage = round(
        (
            missing_values
            / total_cells
        )
        * 100,
        2
    )

    duplicate_percentage = round(
        (
            duplicate_rows
            / max(
                total_rows,
                1
            )
        )
        * 100,
        2
    )

    missing_quality_component = (
        1
        - (
            missing_values
            / total_cells
        )
    )

    duplicate_quality_component = (
        1
        - (
            duplicate_rows
            / max(
                total_rows,
                1
            )
        )
    )

    quality_score = round(
        max(
            0,
            min(
                100,
                (
                    missing_quality_component
                    * 80
                )
                + (
                    duplicate_quality_component
                    * 20
                )
            )
        ),
        2
    )

    if quality_score >= 95:
        quality_status = "Excellent"

    elif quality_score >= 85:
        quality_status = "Good"

    elif quality_score >= 70:
        quality_status = "Fair"

    else:
        quality_status = (
            "Needs Improvement"
        )

    return {
        "total_rows": total_rows,
        "total_columns": total_columns,
        "missing_values": missing_values,
        "missing_percentage": (
            missing_percentage
        ),
        "duplicate_rows": duplicate_rows,
        "duplicate_percentage": (
            duplicate_percentage
        ),
        "numeric_columns": numeric_columns,
        "categorical_columns": (
            categorical_columns
        ),
        "quality_score": quality_score,
        "quality_status": quality_status
    }


# ==========================================================
# Worksheet Styling Helpers
# ==========================================================

def configure_sheet_view(
    worksheet: Worksheet,
    zoom_scale: int = 90
) -> None:
    """
    Configure basic worksheet display settings.
    """

    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = zoom_scale
    worksheet.freeze_panes = None


def style_report_title(
    worksheet: Worksheet,
    title: str,
    subtitle: str,
    end_column: int = 8
) -> None:
    """
    Add a branded title block to the top of a worksheet.
    """

    end_column_letter = get_column_letter(
        end_column
    )

    worksheet.merge_cells(
        f"A1:{end_column_letter}2"
    )

    title_cell = worksheet["A1"]
    title_cell.value = title

    title_cell.font = Font(
        name="Aptos Display",
        size=22,
        bold=True,
        color=WHITE
    )

    title_cell.fill = TITLE_FILL

    title_cell.alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    worksheet.merge_cells(
        f"A3:{end_column_letter}3"
    )

    subtitle_cell = worksheet["A3"]
    subtitle_cell.value = subtitle

    subtitle_cell.font = Font(
        name="Aptos",
        size=10,
        italic=True,
        color=SECONDARY_TEXT
    )

    subtitle_cell.fill = PatternFill(
        fill_type="solid",
        fgColor=LIGHT_BLUE
    )

    subtitle_cell.alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    worksheet.row_dimensions[1].height = 26
    worksheet.row_dimensions[2].height = 26
    worksheet.row_dimensions[3].height = 22


def style_section_heading(
    worksheet: Worksheet,
    row: int,
    title: str,
    start_column: int = 1,
    end_column: int = 8
) -> None:
    """
    Add a blue section heading spanning multiple columns.
    """

    start_letter = get_column_letter(
        start_column
    )

    end_letter = get_column_letter(
        end_column
    )

    worksheet.merge_cells(
        (
            f"{start_letter}{row}:"
            f"{end_letter}{row}"
        )
    )

    cell = worksheet.cell(
        row=row,
        column=start_column
    )

    cell.value = title

    cell.font = Font(
        name="Aptos Display",
        size=13,
        bold=True,
        color=WHITE
    )

    cell.fill = SECTION_FILL

    cell.alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    cell.border = THIN_BORDER

    worksheet.row_dimensions[
        row
    ].height = 23


def style_table_header(
    worksheet: Worksheet,
    row: int,
    start_column: int,
    end_column: int
) -> None:
    """
    Apply consistent formatting to a table header row.
    """

    for column in range(
        start_column,
        end_column + 1
    ):
        cell = worksheet.cell(
            row=row,
            column=column
        )

        cell.font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color=WHITE
        )

        cell.fill = HEADER_FILL

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        cell.border = THIN_BORDER

    worksheet.row_dimensions[
        row
    ].height = 24


def style_data_region(
    worksheet: Worksheet,
    start_row: int,
    end_row: int,
    start_column: int,
    end_column: int
) -> None:
    """
    Format a rectangular data region with alternating rows.
    """

    for row in range(
        start_row,
        end_row + 1
    ):
        for column in range(
            start_column,
            end_column + 1
        ):
            cell = worksheet.cell(
                row=row,
                column=column
            )

            cell.font = Font(
                name="Aptos",
                size=10,
                color=DARK_TEXT
            )

            cell.alignment = Alignment(
                vertical="center",
                wrap_text=False
            )

            cell.border = THIN_BORDER

            if row % 2 == 0:
                cell.fill = (
                    ALTERNATE_ROW_FILL
                )
            else:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=WHITE
                )


def set_reasonable_column_widths(
    worksheet: Worksheet,
    minimum_width: float = 10,
    maximum_width: float = 40,
    padding: float = 2
) -> None:
    """
    Auto-size worksheet columns while preventing extremely wide
    columns from damaging the report layout.
    """

    for column_cells in (
        worksheet.iter_cols()
    ):
        maximum_length = 0

        column_letter = (
            get_column_letter(
                column_cells[0].column
            )
        )

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            value_lines = str(
                value
            ).splitlines()

            cell_length = max(
                len(line)
                for line in value_lines
            )

            maximum_length = max(
                maximum_length,
                cell_length
            )

        calculated_width = min(
            maximum_width,
            max(
                minimum_width,
                maximum_length
                + padding
            )
        )

        worksheet.column_dimensions[
            column_letter
        ].width = calculated_width


def add_report_footer(
    worksheet: Worksheet,
    footer_text: str = (
        "Generated by AI Business "
        "Intelligence Platform"
    )
) -> None:
    """
    Add workbook footer content for printing.
    """

    worksheet.oddFooter.center.text = (
        footer_text
    )

    worksheet.oddFooter.center.size = 8
    worksheet.oddFooter.center.font = (
        "Aptos"
    )

    worksheet.oddFooter.right.text = (
        "Page &P of &N"
    )

    worksheet.oddFooter.right.size = 8
    worksheet.oddFooter.right.font = (
        "Aptos"
    )


def apply_print_settings(
    worksheet: Worksheet,
    repeat_header_row: int | None = None
) -> None:
    """
    Apply print and page-layout settings.
    """

    worksheet.page_setup.orientation = (
        "landscape"
    )

    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0

    worksheet.sheet_properties.pageSetUpPr.fitToPage = (
        True
    )

    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.5
    worksheet.page_margins.bottom = 0.5
    worksheet.page_margins.header = 0.2
    worksheet.page_margins.footer = 0.2

    worksheet.print_options.horizontalCentered = (
        True
    )

    if repeat_header_row:
        worksheet.print_title_rows = (
            f"{repeat_header_row}:"
            f"{repeat_header_row}"
        )


# ==========================================================
# Quality Styling
# ==========================================================

def apply_quality_status_style(
    cell: Cell,
    quality_score: float
) -> None:
    """
    Apply fill and font colors according to the quality score.
    """

    if quality_score >= 85:
        cell.fill = SUCCESS_FILL
        cell.font = Font(
            name="Aptos",
            size=12,
            bold=True,
            color=SUCCESS_GREEN
        )

    elif quality_score >= 70:
        cell.fill = WARNING_FILL
        cell.font = Font(
            name="Aptos",
            size=12,
            bold=True,
            color=WARNING_ORANGE
        )

    else:
        cell.fill = DANGER_FILL
        cell.font = Font(
            name="Aptos",
            size=12,
            bold=True,
            color=DANGER_RED
        )


# ==========================================================
# Sheet 1 — Executive Summary
# ==========================================================

def build_executive_summary_sheet(
    workbook: Workbook,
    df: pd.DataFrame,
    filename: str,
    metrics: dict[str, Any] | None = None
) -> Worksheet:
    """
    Create the Executive Summary worksheet.
    """

    if metrics is None:
        metrics = (
            calculate_dataset_metrics(
                df
            )
        )

    worksheet = workbook.create_sheet(
        title=safe_sheet_name(
            "Executive Summary"
        )
    )

    configure_sheet_view(
        worksheet,
        zoom_scale=95
    )

    generated_at = (
        datetime.now().strftime(
            "%d %B %Y, %I:%M %p"
        )
    )

    style_report_title(
        worksheet=worksheet,
        title=(
            "Executive Business "
            "Intelligence Report"
        ),
        subtitle=(
            f"Dataset: {filename}  |  "
            f"Generated: {generated_at}"
        ),
        end_column=8
    )

    style_section_heading(
        worksheet=worksheet,
        row=5,
        title="Dataset Information",
        start_column=1,
        end_column=8
    )

    information_rows = [
        (
            "Dataset Name",
            filename,
            "Generated Date",
            generated_at
        ),
        (
            "Total Rows",
            metrics[
                "total_rows"
            ],
            "Total Columns",
            metrics[
                "total_columns"
            ]
        ),
        (
            "Numeric Columns",
            metrics[
                "numeric_columns"
            ],
            "Categorical Columns",
            metrics[
                "categorical_columns"
            ]
        )
    ]

    current_row = 6

    for (
        left_label,
        left_value,
        right_label,
        right_value
    ) in information_rows:
        worksheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=2
        )

        worksheet.merge_cells(
            start_row=current_row,
            start_column=3,
            end_row=current_row,
            end_column=4
        )

        worksheet.merge_cells(
            start_row=current_row,
            start_column=5,
            end_row=current_row,
            end_column=6
        )

        worksheet.merge_cells(
            start_row=current_row,
            start_column=7,
            end_row=current_row,
            end_column=8
        )

        left_label_cell = worksheet.cell(
            row=current_row,
            column=1
        )

        left_label_cell.value = left_label

        left_value_cell = worksheet.cell(
            row=current_row,
            column=3
        )

        left_value_cell.value = (
            safe_excel_value(
                left_value
            )
        )

        right_label_cell = worksheet.cell(
            row=current_row,
            column=5
        )

        right_label_cell.value = (
            right_label
        )

        right_value_cell = worksheet.cell(
            row=current_row,
            column=7
        )

        right_value_cell.value = (
            safe_excel_value(
                right_value
            )
        )

        for label_cell in (
            left_label_cell,
            right_label_cell
        ):
            label_cell.font = Font(
                name="Aptos",
                size=10,
                bold=True,
                color=DARK_BLUE
            )

            label_cell.fill = (
                LIGHT_HEADER_FILL
            )

            label_cell.alignment = Alignment(
                horizontal="left",
                vertical="center"
            )

        for value_cell in (
            left_value_cell,
            right_value_cell
        ):
            value_cell.font = Font(
                name="Aptos",
                size=10,
                color=DARK_TEXT
            )

            value_cell.fill = PatternFill(
                fill_type="solid",
                fgColor=WHITE
            )

            value_cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True
            )

        for column in range(
            1,
            9
        ):
            worksheet.cell(
                row=current_row,
                column=column
            ).border = THIN_BORDER

        worksheet.row_dimensions[
            current_row
        ].height = 25

        current_row += 1

    quality_section_row = (
        current_row + 1
    )

    style_section_heading(
        worksheet=worksheet,
        row=quality_section_row,
        title="Data Quality Overview",
        start_column=1,
        end_column=8
    )

    quality_header_row = (
        quality_section_row + 1
    )

    quality_value_row = (
        quality_header_row + 1
    )

    quality_headers = [
        "Missing Values",
        "Missing %",
        "Duplicate Rows",
        "Duplicate %",
        "Quality Score",
        "Quality Status"
    ]

    quality_values = [
        metrics[
            "missing_values"
        ],
        metrics[
            "missing_percentage"
        ],
        metrics[
            "duplicate_rows"
        ],
        metrics[
            "duplicate_percentage"
        ],
        metrics[
            "quality_score"
        ],
        metrics[
            "quality_status"
        ]
    ]

    for index, header in enumerate(
        quality_headers,
        start=1
    ):
        cell = worksheet.cell(
            row=quality_header_row,
            column=index
        )

        cell.value = header

    style_table_header(
        worksheet=worksheet,
        row=quality_header_row,
        start_column=1,
        end_column=len(
            quality_headers
        )
    )

    for index, value in enumerate(
        quality_values,
        start=1
    ):
        cell = worksheet.cell(
            row=quality_value_row,
            column=index
        )

        cell.value = safe_excel_value(
            value
        )

        cell.font = Font(
            name="Aptos",
            size=12,
            bold=True,
            color=DARK_TEXT
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = THIN_BORDER

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=WHITE
        )

    worksheet.cell(
        row=quality_value_row,
        column=2
    ).number_format = '0.00"%"'

    worksheet.cell(
        row=quality_value_row,
        column=4
    ).number_format = '0.00"%"'

    worksheet.cell(
        row=quality_value_row,
        column=5
    ).number_format = '0.00"%"'

    apply_quality_status_style(
        cell=worksheet.cell(
            row=quality_value_row,
            column=5
        ),
        quality_score=metrics[
            "quality_score"
        ]
    )

    apply_quality_status_style(
        cell=worksheet.cell(
            row=quality_value_row,
            column=6
        ),
        quality_score=metrics[
            "quality_score"
        ]
    )

    worksheet.row_dimensions[
        quality_value_row
    ].height = 30

    summary_section_row = (
        quality_value_row + 2
    )

    style_section_heading(
        worksheet=worksheet,
        row=summary_section_row,
        title="Executive Observations",
        start_column=1,
        end_column=8
    )

    observations = [
        (
            f"The dataset contains "
            f"{metrics['total_rows']:,} rows "
            f"and {metrics['total_columns']} columns."
        ),
        (
            f"The overall data-quality score is "
            f"{metrics['quality_score']}%, classified "
            f"as {metrics['quality_status']}."
        ),
        (
            f"A total of "
            f"{metrics['missing_values']:,} missing "
            f"values were identified."
        ),
        (
            f"A total of "
            f"{metrics['duplicate_rows']:,} duplicate "
            f"rows were identified."
        ),
        (
            f"The dataset contains "
            f"{metrics['numeric_columns']} numeric "
            f"columns and "
            f"{metrics['categorical_columns']} "
            f"categorical columns."
        )
    ]

    observation_start_row = (
        summary_section_row + 1
    )

    for offset, observation in enumerate(
        observations
    ):
        row = (
            observation_start_row
            + offset
        )

        worksheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=8
        )

        cell = worksheet.cell(
            row=row,
            column=1
        )

        cell.value = (
            f"• {observation}"
        )

        cell.font = Font(
            name="Aptos",
            size=10,
            color=DARK_TEXT
        )

        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

        cell.border = THIN_BORDER

        if offset % 2 == 0:
            cell.fill = (
                ALTERNATE_ROW_FILL
            )
        else:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=WHITE
            )

        worksheet.row_dimensions[
            row
        ].height = 27

    worksheet.freeze_panes = "A5"

    worksheet.column_dimensions[
        "A"
    ].width = 18

    worksheet.column_dimensions[
        "B"
    ].width = 6

    worksheet.column_dimensions[
        "C"
    ].width = 22

    worksheet.column_dimensions[
        "D"
    ].width = 6

    worksheet.column_dimensions[
        "E"
    ].width = 20

    worksheet.column_dimensions[
        "F"
    ].width = 6

    worksheet.column_dimensions[
        "G"
    ].width = 24

    worksheet.column_dimensions[
        "H"
    ].width = 6

    apply_print_settings(
        worksheet
    )

    add_report_footer(
        worksheet
    )

    worksheet.sheet_properties.tabColor = (
        PRIMARY_BLUE
    )

    return worksheet
# ==========================================================
# DataFrame Writing Helpers
# ==========================================================

def write_dataframe_to_sheet(
    worksheet: Worksheet,
    df: pd.DataFrame,
    start_row: int,
    start_column: int = 1,
    maximum_rows: int | None = None,
    include_index: bool = False,
    add_filter: bool = True,
    freeze_header: bool = True
) -> tuple[int, int]:
    """
    Write a pandas DataFrame into an Excel worksheet.

    Returns:
        A tuple containing:
        - final written row
        - final written column
    """

    if maximum_rows is not None:
        export_df = df.head(
            maximum_rows
        ).copy()
    else:
        export_df = df.copy()

    column_names = [
        str(column)
        for column in export_df.columns
    ]

    if include_index:
        headers = [
            "Index",
            *column_names
        ]
    else:
        headers = column_names

    header_row = start_row

    for offset, header in enumerate(
        headers
    ):
        column_number = (
            start_column
            + offset
        )

        cell = worksheet.cell(
            row=header_row,
            column=column_number
        )

        cell.value = header

    final_column = (
        start_column
        + len(headers)
        - 1
    )

    style_table_header(
        worksheet=worksheet,
        row=header_row,
        start_column=start_column,
        end_column=final_column
    )

    data_start_row = (
        header_row + 1
    )

    for row_offset, (
        index_value,
        row_values
    ) in enumerate(
        export_df.iterrows()
    ):
        excel_row = (
            data_start_row
            + row_offset
        )

        values_to_write = []

        if include_index:
            values_to_write.append(
                index_value
            )

        values_to_write.extend(
            row_values.tolist()
        )

        for column_offset, value in enumerate(
            values_to_write
        ):
            excel_column = (
                start_column
                + column_offset
            )

            cell = worksheet.cell(
                row=excel_row,
                column=excel_column
            )

            cell.value = safe_excel_value(
                value
            )

            cell.font = Font(
                name="Aptos",
                size=10,
                color=DARK_TEXT
            )

            cell.alignment = Alignment(
                vertical="center",
                horizontal="left",
                wrap_text=False
            )

            cell.border = THIN_BORDER

            if row_offset % 2 == 1:
                cell.fill = (
                    ALTERNATE_ROW_FILL
                )
            else:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=WHITE
                )

            apply_excel_number_format(
                cell,
                value
            )

    final_row = (
        header_row
        + max(
            len(export_df),
            0
        )
    )

    if add_filter and headers:
        start_letter = get_column_letter(
            start_column
        )

        end_letter = get_column_letter(
            final_column
        )

        worksheet.auto_filter.ref = (
            f"{start_letter}{header_row}:"
            f"{end_letter}{max(final_row, header_row)}"
        )

    if freeze_header:
        freeze_row = (
            header_row + 1
        )

        freeze_column_letter = (
            get_column_letter(
                start_column
            )
        )

        worksheet.freeze_panes = (
            f"{freeze_column_letter}"
            f"{freeze_row}"
        )

    return (
        final_row,
        final_column
    )


def apply_excel_number_format(
    cell: Cell,
    original_value: Any
) -> None:
    """
    Apply a suitable Excel number format based on the original value.
    """

    if original_value is None:
        return

    try:
        if pd.isna(
            original_value
        ):
            return
    except (
        TypeError,
        ValueError
    ):
        pass

    if isinstance(
        original_value,
        bool
    ):
        cell.number_format = (
            "General"
        )

        return

    if isinstance(
        original_value,
        (
            pd.Timestamp,
            datetime
        )
    ):
        cell.number_format = (
            "dd-mmm-yyyy hh:mm"
        )

        return

    if isinstance(
        original_value,
        int
    ):
        cell.number_format = (
            '#,##0'
        )

        return

    if isinstance(
        original_value,
        float
    ):
        cell.number_format = (
            '#,##0.00'
        )

        return

    if hasattr(
        original_value,
        "dtype"
    ):
        try:
            dtype_name = str(
                original_value.dtype
            ).lower()

            if "int" in dtype_name:
                cell.number_format = (
                    '#,##0'
                )

            elif "float" in dtype_name:
                cell.number_format = (
                    '#,##0.00'
                )

        except (
            TypeError,
            ValueError,
            AttributeError
        ):
            pass


def set_dataframe_column_widths(
    worksheet: Worksheet,
    df: pd.DataFrame,
    start_column: int = 1,
    include_index: bool = False,
    sample_rows: int = 200,
    minimum_width: float = 11,
    maximum_width: float = 34
) -> None:
    """
    Set practical widths based on column names and sample values.
    """

    sampled_df = df.head(
        sample_rows
    )

    offset = 0

    if include_index:
        worksheet.column_dimensions[
            get_column_letter(
                start_column
            )
        ].width = 12

        offset = 1

    for dataframe_column_index, column_name in enumerate(
        sampled_df.columns
    ):
        excel_column = (
            start_column
            + dataframe_column_index
            + offset
        )

        column_letter = get_column_letter(
            excel_column
        )

        maximum_length = len(
            str(column_name)
        )

        for value in sampled_df[
            column_name
        ].tolist():
            if value is None:
                continue

            try:
                if pd.isna(value):
                    continue
            except (
                TypeError,
                ValueError
            ):
                pass

            text_value = str(
                value
            )

            if "\n" in text_value:
                text_length = max(
                    len(line)
                    for line in text_value.splitlines()
                )
            else:
                text_length = len(
                    text_value
                )

            maximum_length = max(
                maximum_length,
                text_length
            )

        width = min(
            maximum_width,
            max(
                minimum_width,
                maximum_length + 2
            )
        )

        worksheet.column_dimensions[
            column_letter
        ].width = width


def add_dataset_metadata_box(
    worksheet: Worksheet,
    filename: str,
    total_rows: int,
    displayed_rows: int,
    start_row: int = 5,
    end_column: int = 8
) -> int:
    """
    Add a dataset information panel above a table.

    Returns the row immediately after the metadata panel.
    """

    style_section_heading(
        worksheet=worksheet,
        row=start_row,
        title="Dataset Information",
        start_column=1,
        end_column=end_column
    )

    metadata = [
        (
            "Dataset Name",
            filename
        ),
        (
            "Total Records",
            total_rows
        ),
        (
            "Displayed Records",
            displayed_rows
        ),
        (
            "Generated At",
            datetime.now().strftime(
                "%d %B %Y, %I:%M %p"
            )
        )
    ]

    metadata_start_row = (
        start_row + 1
    )

    for offset, (
        label,
        value
    ) in enumerate(
        metadata
    ):
        row = (
            metadata_start_row
            + offset
        )

        worksheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=2
        )

        worksheet.merge_cells(
            start_row=row,
            start_column=3,
            end_row=row,
            end_column=end_column
        )

        label_cell = worksheet.cell(
            row=row,
            column=1
        )

        label_cell.value = label

        label_cell.font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color=DARK_BLUE
        )

        label_cell.fill = (
            LIGHT_HEADER_FILL
        )

        label_cell.alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        value_cell = worksheet.cell(
            row=row,
            column=3
        )

        value_cell.value = safe_excel_value(
            value
        )

        value_cell.font = Font(
            name="Aptos",
            size=10,
            color=DARK_TEXT
        )

        value_cell.fill = PatternFill(
            fill_type="solid",
            fgColor=WHITE
        )

        value_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

        for column in range(
            1,
            end_column + 1
        ):
            worksheet.cell(
                row=row,
                column=column
            ).border = THIN_BORDER

        worksheet.row_dimensions[
            row
        ].height = 23

    return (
        metadata_start_row
        + len(metadata)
    )


# ==========================================================
# Sheet 2 — Dataset Preview
# ==========================================================

def build_dataset_preview_sheet(
    workbook: Workbook,
    df: pd.DataFrame,
    filename: str,
    preview_rows: int = 100
) -> Worksheet:
    """
    Create a formatted worksheet containing the first records
    from the uploaded dataset.
    """

    worksheet = workbook.create_sheet(
        title=safe_sheet_name(
            "Dataset Preview"
        )
    )

    configure_sheet_view(
        worksheet=worksheet,
        zoom_scale=85
    )

    displayed_rows = min(
        len(df),
        preview_rows
    )

    style_report_title(
        worksheet=worksheet,
        title="Dataset Preview",
        subtitle=(
            f"First {displayed_rows:,} records "
            f"from {filename}"
        ),
        end_column=max(
            8,
            min(
                len(df.columns),
                16
            )
        )
    )

    metadata_end_row = (
        add_dataset_metadata_box(
            worksheet=worksheet,
            filename=filename,
            total_rows=len(df),
            displayed_rows=displayed_rows,
            start_row=5,
            end_column=max(
                8,
                min(
                    len(df.columns),
                    16
                )
            )
        )
    )

    table_section_row = (
        metadata_end_row + 1
    )

    table_end_column = max(
        1,
        len(df.columns)
    )

    style_section_heading(
        worksheet=worksheet,
        row=table_section_row,
        title="Preview Records",
        start_column=1,
        end_column=table_end_column
    )

    table_header_row = (
        table_section_row + 1
    )

    final_row, final_column = (
        write_dataframe_to_sheet(
            worksheet=worksheet,
            df=df,
            start_row=table_header_row,
            start_column=1,
            maximum_rows=preview_rows,
            include_index=False,
            add_filter=True,
            freeze_header=True
        )
    )

    set_dataframe_column_widths(
        worksheet=worksheet,
        df=df.head(
            preview_rows
        ),
        start_column=1,
        include_index=False,
        sample_rows=preview_rows,
        minimum_width=11,
        maximum_width=32
    )

    worksheet.freeze_panes = (
        f"A{table_header_row + 1}"
    )

    worksheet.sheet_properties.tabColor = (
        "0EA5E9"
    )

    worksheet.auto_filter.ref = (
        f"A{table_header_row}:"
        f"{get_column_letter(final_column)}"
        f"{max(final_row, table_header_row)}"
    )

    apply_print_settings(
        worksheet=worksheet,
        repeat_header_row=table_header_row
    )

    add_report_footer(
        worksheet
    )

    return worksheet
# ==========================================================
# Sheet 3 — Complete Dataset
# ==========================================================

def build_complete_dataset_sheet(
    workbook: Workbook,
    df: pd.DataFrame,
    filename: str
) -> Worksheet:
    """
    Create a worksheet containing the complete uploaded dataset.
    """

    worksheet = workbook.create_sheet(
        title=safe_sheet_name(
            "Complete Dataset"
        )
    )

    configure_sheet_view(
        worksheet=worksheet,
        zoom_scale=80
    )

    title_end_column = max(
        8,
        min(
            len(df.columns),
            16
        )
    )

    style_report_title(
        worksheet=worksheet,
        title="Complete Dataset",
        subtitle=(
            f"Full exported dataset: {filename}"
        ),
        end_column=title_end_column
    )

    metadata_end_row = (
        add_dataset_metadata_box(
            worksheet=worksheet,
            filename=filename,
            total_rows=len(df),
            displayed_rows=len(df),
            start_row=5,
            end_column=title_end_column
        )
    )

    table_section_row = (
        metadata_end_row + 1
    )

    table_end_column = max(
        1,
        len(df.columns)
    )

    style_section_heading(
        worksheet=worksheet,
        row=table_section_row,
        title="All Dataset Records",
        start_column=1,
        end_column=table_end_column
    )

    table_header_row = (
        table_section_row + 1
    )

    final_row, final_column = (
        write_dataframe_to_sheet(
            worksheet=worksheet,
            df=df,
            start_row=table_header_row,
            start_column=1,
            maximum_rows=None,
            include_index=False,
            add_filter=True,
            freeze_header=True
        )
    )

    set_dataframe_column_widths(
        worksheet=worksheet,
        df=df,
        start_column=1,
        include_index=False,
        sample_rows=250,
        minimum_width=11,
        maximum_width=34
    )

    worksheet.freeze_panes = (
        f"A{table_header_row + 1}"
    )

    worksheet.auto_filter.ref = (
        f"A{table_header_row}:"
        f"{get_column_letter(final_column)}"
        f"{max(final_row, table_header_row)}"
    )

    worksheet.sheet_properties.tabColor = (
        "0284C7"
    )

    apply_print_settings(
        worksheet=worksheet,
        repeat_header_row=table_header_row
    )

    add_report_footer(
        worksheet
    )

    return worksheet


# ==========================================================
# Numeric Statistics Helpers
# ==========================================================

def calculate_numeric_statistics(
    df: pd.DataFrame
) -> pd.DataFrame:
    """
    Calculate descriptive statistics for all numeric columns.

    The returned DataFrame contains one row per numeric column.
    """

    numeric_df = df.select_dtypes(
        include="number"
    ).copy()

    statistics_columns = [
        "Column",
        "Count",
        "Missing",
        "Mean",
        "Median",
        "Minimum",
        "Maximum",
        "Std Dev",
        "Variance",
        "25th Percentile",
        "75th Percentile",
        "Unique Values"
    ]

    if numeric_df.empty:
        return pd.DataFrame(
            columns=statistics_columns
        )

    records: list[dict[str, Any]] = []

    for column_name in numeric_df.columns:
        series = pd.to_numeric(
            numeric_df[
                column_name
            ],
            errors="coerce"
        )

        non_null_series = series.dropna()

        count = int(
            non_null_series.count()
        )

        missing = int(
            series.isna().sum()
        )

        if count > 0:
            mean_value = float(
                non_null_series.mean()
            )

            median_value = float(
                non_null_series.median()
            )

            minimum_value = float(
                non_null_series.min()
            )

            maximum_value = float(
                non_null_series.max()
            )

            percentile_25 = float(
                non_null_series.quantile(
                    0.25
                )
            )

            percentile_75 = float(
                non_null_series.quantile(
                    0.75
                )
            )

            unique_values = int(
                non_null_series.nunique()
            )
        else:
            mean_value = None
            median_value = None
            minimum_value = None
            maximum_value = None
            percentile_25 = None
            percentile_75 = None
            unique_values = 0

        if count > 1:
            standard_deviation = float(
                non_null_series.std()
            )

            variance = float(
                non_null_series.var()
            )
        else:
            standard_deviation = None
            variance = None

        records.append(
            {
                "Column": str(
                    column_name
                ),
                "Count": count,
                "Missing": missing,
                "Mean": mean_value,
                "Median": median_value,
                "Minimum": minimum_value,
                "Maximum": maximum_value,
                "Std Dev": standard_deviation,
                "Variance": variance,
                "25th Percentile": percentile_25,
                "75th Percentile": percentile_75,
                "Unique Values": unique_values
            }
        )

    return pd.DataFrame(
        records,
        columns=statistics_columns
    )


def apply_statistics_number_formats(
    worksheet: Worksheet,
    header_row: int,
    start_row: int,
    end_row: int,
    start_column: int = 1
) -> None:
    """
    Apply number formats to the numeric statistics table.
    """

    if end_row < start_row:
        return

    header_map: dict[str, int] = {}

    for cell in worksheet[
        header_row
    ]:
        if cell.value is not None:
            header_map[
                str(
                    cell.value
                )
            ] = cell.column

    integer_headers = {
        "Count",
        "Missing",
        "Unique Values"
    }

    decimal_headers = {
        "Mean",
        "Median",
        "Minimum",
        "Maximum",
        "Std Dev",
        "Variance",
        "25th Percentile",
        "75th Percentile"
    }

    for header in integer_headers:
        column_number = header_map.get(
            header
        )

        if column_number is None:
            continue

        for row in range(
            start_row,
            end_row + 1
        ):
            worksheet.cell(
                row=row,
                column=column_number
            ).number_format = '#,##0'

    for header in decimal_headers:
        column_number = header_map.get(
            header
        )

        if column_number is None:
            continue

        for row in range(
            start_row,
            end_row + 1
        ):
            worksheet.cell(
                row=row,
                column=column_number
            ).number_format = '#,##0.00'


def add_no_numeric_data_message(
    worksheet: Worksheet,
    start_row: int,
    end_column: int = 8
) -> None:
    """
    Add an explanatory message when the dataset has no numeric columns.
    """

    worksheet.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row + 2,
        end_column=end_column
    )

    cell = worksheet.cell(
        row=start_row,
        column=1
    )

    cell.value = (
        "No numeric columns were found in this dataset. "
        "Numeric descriptive statistics could not be generated."
    )

    cell.font = Font(
        name="Aptos",
        size=12,
        bold=True,
        color=WARNING_ORANGE
    )

    cell.fill = WARNING_FILL

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    cell.border = THIN_BORDER

    for row in range(
        start_row,
        start_row + 3
    ):
        worksheet.row_dimensions[
            row
        ].height = 28


# ==========================================================
# Sheet 4 — Numeric Statistics
# ==========================================================

def build_numeric_statistics_sheet(
    workbook: Workbook,
    df: pd.DataFrame,
    filename: str
) -> Worksheet:
    """
    Create a worksheet with descriptive statistics for every
    numeric column in the dataset.
    """

    worksheet = workbook.create_sheet(
        title=safe_sheet_name(
            "Numeric Statistics"
        )
    )

    configure_sheet_view(
        worksheet=worksheet,
        zoom_scale=85
    )

    statistics_df = (
        calculate_numeric_statistics(
            df
        )
    )

    numeric_column_count = int(
        len(
            df.select_dtypes(
                include="number"
            ).columns
        )
    )

    title_end_column = max(
        8,
        min(
            max(
                len(
                    statistics_df.columns
                ),
                1
            ),
            12
        )
    )

    style_report_title(
        worksheet=worksheet,
        title="Numeric Statistics",
        subtitle=(
            f"Descriptive statistical analysis for "
            f"{filename}"
        ),
        end_column=title_end_column
    )

    style_section_heading(
        worksheet=worksheet,
        row=5,
        title="Analysis Overview",
        start_column=1,
        end_column=title_end_column
    )

    overview_rows = [
        (
            "Dataset Name",
            filename
        ),
        (
            "Numeric Columns",
            numeric_column_count
        ),
        (
            "Total Records",
            len(df)
        ),
        (
            "Generated At",
            datetime.now().strftime(
                "%d %B %Y, %I:%M %p"
            )
        )
    ]

    current_row = 6

    for label, value in overview_rows:
        worksheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=2
        )

        worksheet.merge_cells(
            start_row=current_row,
            start_column=3,
            end_row=current_row,
            end_column=title_end_column
        )

        label_cell = worksheet.cell(
            row=current_row,
            column=1
        )

        label_cell.value = label

        label_cell.font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color=DARK_BLUE
        )

        label_cell.fill = (
            LIGHT_HEADER_FILL
        )

        label_cell.alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        value_cell = worksheet.cell(
            row=current_row,
            column=3
        )

        value_cell.value = safe_excel_value(
            value
        )

        value_cell.font = Font(
            name="Aptos",
            size=10,
            color=DARK_TEXT
        )

        value_cell.fill = PatternFill(
            fill_type="solid",
            fgColor=WHITE
        )

        value_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

        for column in range(
            1,
            title_end_column + 1
        ):
            worksheet.cell(
                row=current_row,
                column=column
            ).border = THIN_BORDER

        worksheet.row_dimensions[
            current_row
        ].height = 23

        current_row += 1

    statistics_section_row = (
        current_row + 1
    )

    style_section_heading(
        worksheet=worksheet,
        row=statistics_section_row,
        title="Descriptive Statistics",
        start_column=1,
        end_column=title_end_column
    )

    table_header_row = (
        statistics_section_row + 1
    )

    if statistics_df.empty:
        add_no_numeric_data_message(
            worksheet=worksheet,
            start_row=table_header_row,
            end_column=title_end_column
        )

        worksheet.freeze_panes = (
            f"A{table_header_row}"
        )

    else:
        final_row, final_column = (
            write_dataframe_to_sheet(
                worksheet=worksheet,
                df=statistics_df,
                start_row=table_header_row,
                start_column=1,
                maximum_rows=None,
                include_index=False,
                add_filter=True,
                freeze_header=True
            )
        )

        apply_statistics_number_formats(
            worksheet=worksheet,
            header_row=table_header_row,
            start_row=table_header_row + 1,
            end_row=final_row,
            start_column=1
        )

        set_dataframe_column_widths(
            worksheet=worksheet,
            df=statistics_df,
            start_column=1,
            include_index=False,
            sample_rows=200,
            minimum_width=13,
            maximum_width=24
        )

        worksheet.column_dimensions[
            "A"
        ].width = 24

        worksheet.freeze_panes = (
            f"A{table_header_row + 1}"
        )

        worksheet.auto_filter.ref = (
            f"A{table_header_row}:"
            f"{get_column_letter(final_column)}"
            f"{max(final_row, table_header_row)}"
        )

    worksheet.sheet_properties.tabColor = (
        "7C3AED"
    )

    apply_print_settings(
        worksheet=worksheet,
        repeat_header_row=(
            table_header_row
            if not statistics_df.empty
            else None
        )
    )

    add_report_footer(
        worksheet
    )

    return worksheet
# ==========================================================
# Missing Value Analysis Helpers
# ==========================================================

def calculate_missing_value_report(
    df: pd.DataFrame
) -> pd.DataFrame:
    """
    Create a column-level missing-value analysis report.
    """

    total_rows = max(
        len(df),
        1
    )

    records: list[dict[str, Any]] = []

    for column_name in df.columns:
        series = df[
            column_name
        ]

        missing_count = int(
            series.isna().sum()
        )

        missing_percentage = round(
            (
                missing_count
                / total_rows
            )
            * 100,
            2
        )

        non_missing_count = int(
            series.notna().sum()
        )

        unique_count = int(
            series.nunique(
                dropna=True
            )
        )

        if missing_percentage == 0:
            status = "Complete"

        elif missing_percentage <= 5:
            status = "Low Missing"

        elif missing_percentage <= 20:
            status = "Moderate Missing"

        else:
            status = "High Missing"

        records.append(
            {
                "Column": str(
                    column_name
                ),
                "Data Type": str(
                    series.dtype
                ),
                "Non-Missing Count": (
                    non_missing_count
                ),
                "Missing Count": (
                    missing_count
                ),
                "Missing Percentage": (
                    missing_percentage
                ),
                "Unique Values": (
                    unique_count
                ),
                "Status": status
            }
        )

    report_df = pd.DataFrame(
        records,
        columns=[
            "Column",
            "Data Type",
            "Non-Missing Count",
            "Missing Count",
            "Missing Percentage",
            "Unique Values",
            "Status"
        ]
    )

    if not report_df.empty:
        report_df = report_df.sort_values(
            by=[
                "Missing Percentage",
                "Missing Count"
            ],
            ascending=[
                False,
                False
            ]
        ).reset_index(
            drop=True
        )

    return report_df


def apply_missing_status_style(
    cell: Cell,
    status: str
) -> None:
    """
    Apply status-based formatting to missing-value report cells.
    """

    normalized_status = str(
        status
    ).strip().lower()

    if normalized_status == "complete":
        cell.fill = SUCCESS_FILL

        cell.font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color=SUCCESS_GREEN
        )

    elif normalized_status == "low missing":
        cell.fill = LIGHT_HEADER_FILL

        cell.font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color=DARK_BLUE
        )

    elif normalized_status == "moderate missing":
        cell.fill = WARNING_FILL

        cell.font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color=WARNING_ORANGE
        )

    else:
        cell.fill = DANGER_FILL

        cell.font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color=DANGER_RED
        )

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    cell.border = THIN_BORDER


def apply_missing_report_formats(
    worksheet: Worksheet,
    header_row: int,
    start_row: int,
    end_row: int
) -> None:
    """
    Apply number formats and status styles to a missing-value table.
    """

    if end_row < start_row:
        return

    header_map: dict[str, int] = {}

    for cell in worksheet[
        header_row
    ]:
        if cell.value is not None:
            header_map[
                str(
                    cell.value
                )
            ] = cell.column

    integer_columns = {
        "Non-Missing Count",
        "Missing Count",
        "Unique Values"
    }

    for header in integer_columns:
        column_number = header_map.get(
            header
        )

        if column_number is None:
            continue

        for row in range(
            start_row,
            end_row + 1
        ):
            worksheet.cell(
                row=row,
                column=column_number
            ).number_format = '#,##0'

    percentage_column = header_map.get(
        "Missing Percentage"
    )

    if percentage_column is not None:
        for row in range(
            start_row,
            end_row + 1
        ):
            worksheet.cell(
                row=row,
                column=percentage_column
            ).number_format = '0.00"%"'

    status_column = header_map.get(
        "Status"
    )

    if status_column is not None:
        for row in range(
            start_row,
            end_row + 1
        ):
            status_cell = worksheet.cell(
                row=row,
                column=status_column
            )

            apply_missing_status_style(
                cell=status_cell,
                status=str(
                    status_cell.value
                    or ""
                )
            )


def build_missing_summary_points(
    df: pd.DataFrame,
    report_df: pd.DataFrame
) -> list[str]:
    """
    Build concise observations for the missing-value report.
    """

    total_missing = int(
        df.isna()
        .sum()
        .sum()
    )

    total_columns = int(
        len(df.columns)
    )

    columns_with_missing = int(
        (
            report_df[
                "Missing Count"
            ]
            > 0
        ).sum()
    ) if not report_df.empty else 0

    high_missing_columns = int(
        (
            report_df[
                "Missing Percentage"
            ]
            > 20
        ).sum()
    ) if not report_df.empty else 0

    complete_columns = (
        total_columns
        - columns_with_missing
    )

    points = [
        (
            f"The dataset contains "
            f"{total_missing:,} missing values."
        ),
        (
            f"{columns_with_missing} of "
            f"{total_columns} columns contain "
            f"at least one missing value."
        ),
        (
            f"{complete_columns} columns are fully complete."
        )
    ]

    if high_missing_columns > 0:
        points.append(
            (
                f"{high_missing_columns} columns have more "
                f"than 20% missing values and should be "
                f"reviewed before analysis."
            )
        )
    else:
        points.append(
            (
                "No column exceeds the 20% high-missing "
                "threshold."
            )
        )

    if total_missing == 0:
        points.append(
            (
                "The dataset has no missing values and "
                "requires no missing-data treatment."
            )
        )
    else:
        points.append(
            (
                "Consider imputation, row removal or source "
                "validation depending on the affected fields."
            )
        )

    return points


# ==========================================================
# Sheet 5 — Missing Value Analysis
# ==========================================================

def build_missing_value_sheet(
    workbook: Workbook,
    df: pd.DataFrame,
    filename: str
) -> Worksheet:
    """
    Create a professional missing-value analysis worksheet.
    """

    worksheet = workbook.create_sheet(
        title=safe_sheet_name(
            "Missing Value Analysis"
        )
    )

    configure_sheet_view(
        worksheet=worksheet,
        zoom_scale=90
    )

    report_df = (
        calculate_missing_value_report(
            df
        )
    )

    title_end_column = max(
        8,
        len(
            report_df.columns
        )
        if not report_df.empty
        else 8
    )

    style_report_title(
        worksheet=worksheet,
        title="Missing Value Analysis",
        subtitle=(
            f"Column-level completeness assessment "
            f"for {filename}"
        ),
        end_column=title_end_column
    )

    style_section_heading(
        worksheet=worksheet,
        row=5,
        title="Missing Data Overview",
        start_column=1,
        end_column=title_end_column
    )

    summary_points = (
        build_missing_summary_points(
            df=df,
            report_df=report_df
        )
    )

    summary_start_row = 6

    for offset, point in enumerate(
        summary_points
    ):
        row = (
            summary_start_row
            + offset
        )

        worksheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=title_end_column
        )

        cell = worksheet.cell(
            row=row,
            column=1
        )

        cell.value = f"• {point}"

        cell.font = Font(
            name="Aptos",
            size=10,
            color=DARK_TEXT
        )

        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

        cell.border = THIN_BORDER

        if offset % 2 == 0:
            cell.fill = (
                ALTERNATE_ROW_FILL
            )
        else:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=WHITE
            )

        worksheet.row_dimensions[
            row
        ].height = 27

    report_section_row = (
        summary_start_row
        + len(summary_points)
        + 1
    )

    style_section_heading(
        worksheet=worksheet,
        row=report_section_row,
        title="Column Completeness Report",
        start_column=1,
        end_column=title_end_column
    )

    table_header_row = (
        report_section_row + 1
    )

    if report_df.empty:
        worksheet.merge_cells(
            start_row=table_header_row,
            start_column=1,
            end_row=table_header_row + 2,
            end_column=title_end_column
        )

        message_cell = worksheet.cell(
            row=table_header_row,
            column=1
        )

        message_cell.value = (
            "No columns were found in the uploaded dataset."
        )

        message_cell.font = Font(
            name="Aptos",
            size=12,
            bold=True,
            color=WARNING_ORANGE
        )

        message_cell.fill = WARNING_FILL

        message_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        message_cell.border = THIN_BORDER

    else:
        final_row, final_column = (
            write_dataframe_to_sheet(
                worksheet=worksheet,
                df=report_df,
                start_row=table_header_row,
                start_column=1,
                maximum_rows=None,
                include_index=False,
                add_filter=True,
                freeze_header=True
            )
        )

        apply_missing_report_formats(
            worksheet=worksheet,
            header_row=table_header_row,
            start_row=table_header_row + 1,
            end_row=final_row
        )

        set_dataframe_column_widths(
            worksheet=worksheet,
            df=report_df,
            start_column=1,
            include_index=False,
            sample_rows=250,
            minimum_width=14,
            maximum_width=30
        )

        worksheet.column_dimensions[
            "A"
        ].width = 26

        worksheet.column_dimensions[
            "B"
        ].width = 18

        worksheet.column_dimensions[
            "G"
        ].width = 20

        worksheet.freeze_panes = (
            f"A{table_header_row + 1}"
        )

        worksheet.auto_filter.ref = (
            f"A{table_header_row}:"
            f"{get_column_letter(final_column)}"
            f"{max(final_row, table_header_row)}"
        )

    worksheet.sheet_properties.tabColor = (
        "F59E0B"
    )

    apply_print_settings(
        worksheet=worksheet,
        repeat_header_row=(
            table_header_row
            if not report_df.empty
            else None
        )
    )

    add_report_footer(
        worksheet
    )

    return worksheet


# ==========================================================
# AI Content Helpers
# ==========================================================

def normalize_excel_ai_items(
    items: Any
) -> list[str]:
    """
    Convert AI insights or recommendations into clean strings.
    """

    if items is None:
        return []

    if isinstance(
        items,
        str
    ):
        items = [
            items
        ]

    if isinstance(
        items,
        dict
    ):
        items = [
            items
        ]

    if not isinstance(
        items,
        (
            list,
            tuple,
            set
        )
    ):
        items = [
            items
        ]

    normalized_items: list[str] = []

    preferred_keys = (
        "insight",
        "recommendation",
        "title",
        "message",
        "description",
        "text",
        "content"
    )

    for item in items:
        if item is None:
            continue

        text = ""

        if isinstance(
            item,
            str
        ):
            text = item

        elif isinstance(
            item,
            dict
        ):
            for key in preferred_keys:
                value = item.get(
                    key
                )

                if value:
                    text = str(
                        value
                    )

                    break

            if not text:
                text = " | ".join(
                    (
                        f"{key}: {value}"
                        for key, value
                        in item.items()
                        if value is not None
                    )
                )

        else:
            text = str(
                item
            )

        cleaned_text = " ".join(
            text.split()
        ).strip()

        if cleaned_text:
            normalized_items.append(
                cleaned_text
            )

    return normalized_items


def build_default_ai_insights(
    df: pd.DataFrame
) -> list[str]:
    """
    Build dataset-driven fallback insights when AI output is absent.
    """

    metrics = (
        calculate_dataset_metrics(
            df
        )
    )

    insights = [
        (
            f"The dataset includes "
            f"{metrics['total_rows']:,} records across "
            f"{metrics['total_columns']} columns."
        ),
        (
            f"The overall data-quality score is "
            f"{metrics['quality_score']}%, classified "
            f"as {metrics['quality_status']}."
        ),
        (
            f"{metrics['numeric_columns']} columns are numeric "
            f"and {metrics['categorical_columns']} are "
            f"categorical or non-numeric."
        )
    ]

    if metrics[
        "missing_values"
    ] > 0:
        insights.append(
            (
                f"The dataset contains "
                f"{metrics['missing_values']:,} missing "
                f"values, representing "
                f"{metrics['missing_percentage']}% of "
                f"all cells."
            )
        )
    else:
        insights.append(
            (
                "No missing values were detected in the dataset."
            )
        )

    if metrics[
        "duplicate_rows"
    ] > 0:
        insights.append(
            (
                f"{metrics['duplicate_rows']:,} duplicate rows "
                f"were found and may affect aggregate analysis."
            )
        )
    else:
        insights.append(
            (
                "No duplicate rows were detected."
            )
        )

    numeric_df = df.select_dtypes(
        include="number"
    )

    if not numeric_df.empty:
        highest_variability_column = None
        highest_standard_deviation = None

        for column_name in numeric_df.columns:
            standard_deviation = (
                pd.to_numeric(
                    numeric_df[
                        column_name
                    ],
                    errors="coerce"
                )
                .std()
            )

            if pd.isna(
                standard_deviation
            ):
                continue

            if (
                highest_standard_deviation
                is None
                or standard_deviation
                > highest_standard_deviation
            ):
                highest_standard_deviation = (
                    float(
                        standard_deviation
                    )
                )

                highest_variability_column = (
                    str(
                        column_name
                    )
                )

        if highest_variability_column:
            insights.append(
                (
                    f"'{highest_variability_column}' shows the "
                    f"highest numeric variability and may warrant "
                    f"closer investigation."
                )
            )

    return insights


def add_ai_content_rows(
    worksheet: Worksheet,
    items: list[str],
    start_row: int,
    end_column: int = 8
) -> int:
    """
    Write AI-generated content as professionally formatted rows.

    Returns the final written row.
    """

    if not items:
        items = [
            (
                "No AI-generated content was available "
                "for this report."
            )
        ]

    current_row = start_row

    for index, item in enumerate(
        items,
        start=1
    ):
        worksheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=1
        )

        worksheet.merge_cells(
            start_row=current_row,
            start_column=2,
            end_row=current_row,
            end_column=end_column
        )

        number_cell = worksheet.cell(
            row=current_row,
            column=1
        )

        number_cell.value = index

        number_cell.font = Font(
            name="Aptos",
            size=12,
            bold=True,
            color=WHITE
        )

        number_cell.fill = SECTION_FILL

        number_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        number_cell.border = THIN_BORDER

        content_cell = worksheet.cell(
            row=current_row,
            column=2
        )

        content_cell.value = item

        content_cell.font = Font(
            name="Aptos",
            size=11,
            color=DARK_TEXT
        )

        content_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

        content_cell.border = THIN_BORDER

        if index % 2 == 0:
            content_cell.fill = (
                ALTERNATE_ROW_FILL
            )
        else:
            content_cell.fill = PatternFill(
                fill_type="solid",
                fgColor=WHITE
            )

        estimated_lines = max(
            1,
            len(item)
            // 105
            + 1
        )

        worksheet.row_dimensions[
            current_row
        ].height = max(
            30,
            estimated_lines * 18
        )

        current_row += 1

    return current_row - 1


# ==========================================================
# Sheet 6 — AI Insights
# ==========================================================

def build_ai_insights_sheet(
    workbook: Workbook,
    df: pd.DataFrame,
    filename: str,
    insights: Any = None
) -> Worksheet:
    """
    Create a professional AI Insights worksheet.
    """

    worksheet = workbook.create_sheet(
        title=safe_sheet_name(
            "AI Insights"
        )
    )

    configure_sheet_view(
        worksheet=worksheet,
        zoom_scale=95
    )

    normalized_insights = (
        normalize_excel_ai_items(
            insights
        )
    )

    if not normalized_insights:
        normalized_insights = (
            build_default_ai_insights(
                df
            )
        )

    style_report_title(
        worksheet=worksheet,
        title="AI-Powered Business Insights",
        subtitle=(
            f"Automated observations generated from "
            f"{filename}"
        ),
        end_column=8
    )

    style_section_heading(
        worksheet=worksheet,
        row=5,
        title="Executive AI Observations",
        start_column=1,
        end_column=8
    )

    final_content_row = (
        add_ai_content_rows(
            worksheet=worksheet,
            items=normalized_insights,
            start_row=6,
            end_column=8
        )
    )

    note_row = (
        final_content_row + 2
    )

    worksheet.merge_cells(
        start_row=note_row,
        start_column=1,
        end_row=note_row,
        end_column=8
    )

    note_cell = worksheet.cell(
        row=note_row,
        column=1
    )

    note_cell.value = (
        "Note: AI-generated insights should be reviewed "
        "alongside business context before final decisions."
    )

    note_cell.font = Font(
        name="Aptos",
        size=9,
        italic=True,
        color=MUTED_TEXT
    )

    note_cell.fill = LIGHT_HEADER_FILL

    note_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True
    )

    note_cell.border = THIN_BORDER

    worksheet.row_dimensions[
        note_row
    ].height = 28

    worksheet.column_dimensions[
        "A"
    ].width = 8

    worksheet.column_dimensions[
        "B"
    ].width = 20

    for column in (
        "C",
        "D",
        "E",
        "F",
        "G",
        "H"
    ):
        worksheet.column_dimensions[
            column
        ].width = 15

    worksheet.freeze_panes = "A6"

    worksheet.sheet_properties.tabColor = (
        "8B5CF6"
    )

    apply_print_settings(
        worksheet
    )

    add_report_footer(
        worksheet
    )

    return worksheet
# ==========================================================
# Recommendation Helpers
# ==========================================================

def build_default_recommendations(
    df: pd.DataFrame
) -> list[str]:
    """
    Generate practical dataset-driven recommendations when no
    AI recommendations are supplied.
    """

    metrics = calculate_dataset_metrics(
        df
    )

    recommendations: list[str] = []

    if metrics[
        "missing_values"
    ] > 0:
        recommendations.append(
            (
                "Review columns containing missing values and "
                "apply an appropriate treatment strategy, such "
                "as source correction, imputation or controlled "
                "record exclusion."
            )
        )
    else:
        recommendations.append(
            (
                "Maintain the current data-completeness standard "
                "through automated validation checks during future "
                "data collection and ingestion."
            )
        )

    if metrics[
        "duplicate_rows"
    ] > 0:
        recommendations.append(
            (
                f"Investigate and remove the "
                f"{metrics['duplicate_rows']:,} duplicate rows "
                f"before calculating business totals, averages "
                f"or forecasting models."
            )
        )
    else:
        recommendations.append(
            (
                "Continue enforcing unique-record controls to "
                "prevent duplicate data from entering downstream "
                "analytics and reporting workflows."
            )
        )

    if metrics[
        "quality_score"
    ] < 70:
        recommendations.append(
            (
                "Prioritize a structured data-cleaning initiative "
                "because the current quality score indicates that "
                "the dataset may not yet be suitable for critical "
                "business decisions."
            )
        )

    elif metrics[
        "quality_score"
    ] < 85:
        recommendations.append(
            (
                "Improve the dataset through targeted validation "
                "and standardization before using it for high-impact "
                "executive or predictive analysis."
            )
        )

    else:
        recommendations.append(
            (
                "The dataset is generally suitable for analysis, "
                "but key findings should still be validated against "
                "business definitions and source-system rules."
            )
        )

    numeric_df = df.select_dtypes(
        include="number"
    )

    if not numeric_df.empty:
        recommendations.append(
            (
                "Use the numeric fields to develop trend analysis, "
                "KPI monitoring, anomaly detection and forecasting "
                "models for forward-looking decision support."
            )
        )

    categorical_columns = list(
        df.select_dtypes(
            exclude="number"
        ).columns
    )

    if categorical_columns:
        recommendations.append(
            (
                "Use categorical dimensions to segment performance "
                "by customer, product, location, department or other "
                "relevant business groups."
            )
        )

    recommendations.extend(
        [
            (
                "Create automated data-quality rules for missing "
                "values, duplicate records, invalid formats and "
                "unexpected value ranges."
            ),
            (
                "Define a consistent reporting cadence and monitor "
                "the most important KPIs through the Executive "
                "Dashboard."
            ),
            (
                "Document metric definitions and calculation logic "
                "so all stakeholders interpret the report in the "
                "same way."
            ),
            (
                "Validate AI-generated conclusions with subject "
                "matter experts before taking operational or "
                "financial action."
            )
        ]
    )

    unique_recommendations: list[str] = []

    for recommendation in recommendations:
        cleaned_recommendation = " ".join(
            recommendation.split()
        ).strip()

        if (
            cleaned_recommendation
            and cleaned_recommendation
            not in unique_recommendations
        ):
            unique_recommendations.append(
                cleaned_recommendation
            )

    return unique_recommendations


def classify_recommendation_priority(
    recommendation: str
) -> str:
    """
    Assign a practical priority level based on recommendation text.
    """

    normalized_text = str(
        recommendation
    ).lower()

    high_priority_keywords = (
        "missing",
        "duplicate",
        "invalid",
        "cleaning",
        "quality score",
        "critical",
        "risk",
        "error",
        "correct",
        "remove"
    )

    medium_priority_keywords = (
        "validate",
        "monitor",
        "standard",
        "automated",
        "forecast",
        "anomaly",
        "review",
        "definition",
        "control"
    )

    if any(
        keyword in normalized_text
        for keyword in high_priority_keywords
    ):
        return "High"

    if any(
        keyword in normalized_text
        for keyword in medium_priority_keywords
    ):
        return "Medium"

    return "Normal"


def classify_recommendation_category(
    recommendation: str
) -> str:
    """
    Assign each recommendation to a business category.
    """

    normalized_text = str(
        recommendation
    ).lower()

    if any(
        keyword in normalized_text
        for keyword in (
            "missing",
            "duplicate",
            "quality",
            "clean",
            "validation",
            "format",
            "record",
            "source"
        )
    ):
        return "Data Quality"

    if any(
        keyword in normalized_text
        for keyword in (
            "forecast",
            "predict",
            "trend",
            "anomaly",
            "model",
            "numeric"
        )
    ):
        return "Advanced Analytics"

    if any(
        keyword in normalized_text
        for keyword in (
            "dashboard",
            "kpi",
            "report",
            "monitor",
            "cadence"
        )
    ):
        return "Reporting"

    if any(
        keyword in normalized_text
        for keyword in (
            "segment",
            "customer",
            "product",
            "location",
            "department",
            "category"
        )
    ):
        return "Segmentation"

    if any(
        keyword in normalized_text
        for keyword in (
            "document",
            "definition",
            "stakeholder",
            "expert",
            "governance"
        )
    ):
        return "Governance"

    return "Business Improvement"


def build_recommendation_dataframe(
    recommendations: Any,
    df: pd.DataFrame
) -> pd.DataFrame:
    """
    Normalize recommendations and prepare them for Excel export.
    """

    normalized_recommendations = (
        normalize_excel_ai_items(
            recommendations
        )
    )

    if not normalized_recommendations:
        normalized_recommendations = (
            build_default_recommendations(
                df
            )
        )

    records: list[dict[str, Any]] = []

    for index, recommendation in enumerate(
        normalized_recommendations,
        start=1
    ):
        records.append(
            {
                "No.": index,
                "Priority": (
                    classify_recommendation_priority(
                        recommendation
                    )
                ),
                "Category": (
                    classify_recommendation_category(
                        recommendation
                    )
                ),
                "Recommendation": recommendation,
                "Status": "Pending"
            }
        )

    return pd.DataFrame(
        records,
        columns=[
            "No.",
            "Priority",
            "Category",
            "Recommendation",
            "Status"
        ]
    )


def apply_recommendation_priority_style(
    cell: Cell,
    priority: str
) -> None:
    """
    Apply color formatting according to recommendation priority.
    """

    normalized_priority = str(
        priority
    ).strip().lower()

    if normalized_priority == "high":
        cell.fill = DANGER_FILL

        cell.font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color=DANGER_RED
        )

    elif normalized_priority == "medium":
        cell.fill = WARNING_FILL

        cell.font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color=WARNING_ORANGE
        )

    else:
        cell.fill = SUCCESS_FILL

        cell.font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color=SUCCESS_GREEN
        )

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    cell.border = THIN_BORDER


def apply_recommendation_status_style(
    cell: Cell,
    status: str
) -> None:
    """
    Format recommendation execution-status cells.
    """

    normalized_status = str(
        status
    ).strip().lower()

    if normalized_status in (
        "completed",
        "done",
        "closed"
    ):
        cell.fill = SUCCESS_FILL

        cell.font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color=SUCCESS_GREEN
        )

    elif normalized_status in (
        "in progress",
        "working",
        "ongoing"
    ):
        cell.fill = WARNING_FILL

        cell.font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color=WARNING_ORANGE
        )

    else:
        cell.fill = LIGHT_HEADER_FILL

        cell.font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color=DARK_BLUE
        )

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    cell.border = THIN_BORDER


def apply_recommendation_table_styles(
    worksheet: Worksheet,
    header_row: int,
    start_row: int,
    end_row: int
) -> None:
    """
    Apply specialized formatting to the recommendation table.
    """

    if end_row < start_row:
        return

    header_map: dict[str, int] = {}

    for cell in worksheet[
        header_row
    ]:
        if cell.value is not None:
            header_map[
                str(
                    cell.value
                )
            ] = cell.column

    number_column = header_map.get(
        "No."
    )

    priority_column = header_map.get(
        "Priority"
    )

    recommendation_column = header_map.get(
        "Recommendation"
    )

    status_column = header_map.get(
        "Status"
    )

    for row in range(
        start_row,
        end_row + 1
    ):
        if number_column is not None:
            number_cell = worksheet.cell(
                row=row,
                column=number_column
            )

            number_cell.number_format = (
                '0'
            )

            number_cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        if priority_column is not None:
            priority_cell = worksheet.cell(
                row=row,
                column=priority_column
            )

            apply_recommendation_priority_style(
                cell=priority_cell,
                priority=str(
                    priority_cell.value
                    or ""
                )
            )

        if recommendation_column is not None:
            recommendation_cell = worksheet.cell(
                row=row,
                column=recommendation_column
            )

            recommendation_cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True
            )

            recommendation_text = str(
                recommendation_cell.value
                or ""
            )

            estimated_lines = max(
                2,
                len(
                    recommendation_text
                )
                // 85
                + 1
            )

            worksheet.row_dimensions[
                row
            ].height = max(
                36,
                estimated_lines * 17
            )

        if status_column is not None:
            status_cell = worksheet.cell(
                row=row,
                column=status_column
            )

            apply_recommendation_status_style(
                cell=status_cell,
                status=str(
                    status_cell.value
                    or ""
                )
            )


def build_recommendation_summary(
    recommendation_df: pd.DataFrame
) -> list[str]:
    """
    Create a concise summary of recommendation priorities.
    """

    if recommendation_df.empty:
        return [
            (
                "No recommendations were generated "
                "for this dataset."
            )
        ]

    total_recommendations = int(
        len(
            recommendation_df
        )
    )

    high_priority = int(
        (
            recommendation_df[
                "Priority"
            ]
            == "High"
        ).sum()
    )

    medium_priority = int(
        (
            recommendation_df[
                "Priority"
            ]
            == "Medium"
        ).sum()
    )

    normal_priority = (
        total_recommendations
        - high_priority
        - medium_priority
    )

    categories = int(
        recommendation_df[
            "Category"
        ].nunique()
    )

    return [
        (
            f"{total_recommendations} business recommendations "
            f"were generated from the dataset analysis."
        ),
        (
            f"{high_priority} recommendations are marked as "
            f"high priority."
        ),
        (
            f"{medium_priority} recommendations are marked as "
            f"medium priority."
        ),
        (
            f"{normal_priority} recommendations are classified "
            f"as normal priority."
        ),
        (
            f"The recommendations cover {categories} business "
            f"improvement categories."
        )
    ]


# ==========================================================
# Sheet 7 — Business Recommendations
# ==========================================================

def build_recommendations_sheet(
    workbook: Workbook,
    df: pd.DataFrame,
    filename: str,
    recommendations: Any = None
) -> Worksheet:
    """
    Create a management-ready business recommendations worksheet.
    """

    worksheet = workbook.create_sheet(
        title=safe_sheet_name(
            "Recommendations"
        )
    )

    configure_sheet_view(
        worksheet=worksheet,
        zoom_scale=90
    )

    recommendation_df = (
        build_recommendation_dataframe(
            recommendations=recommendations,
            df=df
        )
    )

    title_end_column = max(
        8,
        len(
            recommendation_df.columns
        )
        if not recommendation_df.empty
        else 8
    )

    style_report_title(
        worksheet=worksheet,
        title="Business Recommendations",
        subtitle=(
            f"Prioritized actions generated from "
            f"the analysis of {filename}"
        ),
        end_column=title_end_column
    )

    style_section_heading(
        worksheet=worksheet,
        row=5,
        title="Recommendation Overview",
        start_column=1,
        end_column=title_end_column
    )

    summary_points = (
        build_recommendation_summary(
            recommendation_df
        )
    )

    summary_start_row = 6

    for offset, point in enumerate(
        summary_points
    ):
        row = (
            summary_start_row
            + offset
        )

        worksheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=title_end_column
        )

        cell = worksheet.cell(
            row=row,
            column=1
        )

        cell.value = f"• {point}"

        cell.font = Font(
            name="Aptos",
            size=10,
            color=DARK_TEXT
        )

        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

        cell.border = THIN_BORDER

        if offset % 2 == 0:
            cell.fill = (
                ALTERNATE_ROW_FILL
            )
        else:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=WHITE
            )

        worksheet.row_dimensions[
            row
        ].height = 27

    table_section_row = (
        summary_start_row
        + len(
            summary_points
        )
        + 1
    )

    style_section_heading(
        worksheet=worksheet,
        row=table_section_row,
        title="Prioritized Action Plan",
        start_column=1,
        end_column=title_end_column
    )

    table_header_row = (
        table_section_row + 1
    )

    if recommendation_df.empty:
        worksheet.merge_cells(
            start_row=table_header_row,
            start_column=1,
            end_row=table_header_row + 2,
            end_column=title_end_column
        )

        empty_cell = worksheet.cell(
            row=table_header_row,
            column=1
        )

        empty_cell.value = (
            "No business recommendations are available."
        )

        empty_cell.font = Font(
            name="Aptos",
            size=12,
            bold=True,
            color=WARNING_ORANGE
        )

        empty_cell.fill = WARNING_FILL

        empty_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        empty_cell.border = THIN_BORDER

    else:
        final_row, final_column = (
            write_dataframe_to_sheet(
                worksheet=worksheet,
                df=recommendation_df,
                start_row=table_header_row,
                start_column=1,
                maximum_rows=None,
                include_index=False,
                add_filter=True,
                freeze_header=True
            )
        )

        apply_recommendation_table_styles(
            worksheet=worksheet,
            header_row=table_header_row,
            start_row=table_header_row + 1,
            end_row=final_row
        )

        worksheet.column_dimensions[
            "A"
        ].width = 8

        worksheet.column_dimensions[
            "B"
        ].width = 14

        worksheet.column_dimensions[
            "C"
        ].width = 22

        worksheet.column_dimensions[
            "D"
        ].width = 70

        worksheet.column_dimensions[
            "E"
        ].width = 16

        for column_number in range(
            6,
            final_column + 1
        ):
            worksheet.column_dimensions[
                get_column_letter(
                    column_number
                )
            ].width = 15

        worksheet.freeze_panes = (
            f"A{table_header_row + 1}"
        )

        worksheet.auto_filter.ref = (
            f"A{table_header_row}:"
            f"{get_column_letter(final_column)}"
            f"{max(final_row, table_header_row)}"
        )

    note_row = (
        (
            final_row + 2
        )
        if not recommendation_df.empty
        else (
            table_header_row + 4
        )
    )

    worksheet.merge_cells(
        start_row=note_row,
        start_column=1,
        end_row=note_row,
        end_column=title_end_column
    )

    note_cell = worksheet.cell(
        row=note_row,
        column=1
    )

    note_cell.value = (
        "Implementation note: Assign an owner and target date "
        "to each recommendation before execution. Update the "
        "Status column as actions progress."
    )

    note_cell.font = Font(
        name="Aptos",
        size=9,
        italic=True,
        color=MUTED_TEXT
    )

    note_cell.fill = LIGHT_HEADER_FILL

    note_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True
    )

    note_cell.border = THIN_BORDER

    worksheet.row_dimensions[
        note_row
    ].height = 30

    worksheet.sheet_properties.tabColor = (
        SUCCESS_GREEN
    )

    apply_print_settings(
        worksheet=worksheet,
        repeat_header_row=(
            table_header_row
            if not recommendation_df.empty
            else None
        )
    )

    add_report_footer(
        worksheet
    )

    return worksheet
# ==========================================================
# Final Professional Excel Generator
# ==========================================================

def generate_professional_excel(
    df: pd.DataFrame,
    filename: str,
    stats: dict[str, Any] | None = None,
    insights: Any = None,
    recommendations: Any = None,
    output_path: str | None = None
) -> str:
    """
    Generate the complete professional Excel business report.

    Workbook sheets:
    1. Executive Summary
    2. Dataset Preview
    3. Complete Dataset
    4. Numeric Statistics
    5. Missing Value Analysis
    6. AI Insights
    7. Recommendations

    Parameters
    ----------
    df:
        Uploaded dataset as a pandas DataFrame.

    filename:
        Original dataset filename.

    stats:
        Optional pre-calculated dataset statistics. The function
        accepts this value for compatibility with the export route.

    insights:
        AI-generated insights supplied as strings, dictionaries
        or a collection of values.

    recommendations:
        AI-generated recommendations supplied as strings,
        dictionaries or a collection of values.

    output_path:
        Full file path where the generated workbook will be saved.

    Returns
    -------
    str
        The final saved Excel workbook path.
    """

    if not isinstance(
        df,
        pd.DataFrame
    ):
        raise TypeError(
            "df must be a pandas DataFrame."
        )

    if len(
        df.columns
    ) == 0:
        raise ValueError(
            (
                "Cannot generate a professional Excel report "
                "because the dataset has no columns."
            )
        )

    if not filename:
        filename = "uploaded_dataset"

    safe_filename = os.path.basename(
        str(
            filename
        )
    )

    dataset_name = os.path.splitext(
        safe_filename
    )[0]

    if not dataset_name:
        dataset_name = (
            "uploaded_dataset"
        )

    if not output_path:
        output_path = os.path.join(
            "exports",
            (
                f"{dataset_name}_"
                f"professional_report.xlsx"
            )
        )

    if not str(
        output_path
    ).lower().endswith(
        ".xlsx"
    ):
        output_path = (
            f"{output_path}.xlsx"
        )

    ensure_output_directory(
        output_path
    )

    calculated_metrics = (
        calculate_dataset_metrics(
            df
        )
    )

    if isinstance(
        stats,
        dict
    ):
        metric_key_map = {
            "rows": "total_rows",
            "total_rows": "total_rows",
            "columns": "total_columns",
            "total_columns": "total_columns",
            "missing": "missing_values",
            "missing_values": "missing_values",
            "duplicates": "duplicate_rows",
            "duplicate_rows": "duplicate_rows",
            "numeric_columns": "numeric_columns",
            "categorical_columns": (
                "categorical_columns"
            ),
            "quality_score": "quality_score",
            "quality_status": (
                "quality_status"
            )
        }

        for source_key, target_key in (
            metric_key_map.items()
        ):
            source_value = stats.get(
                source_key
            )

            if source_value is not None:
                calculated_metrics[
                    target_key
                ] = source_value

    normalized_insights = (
        normalize_excel_ai_items(
            insights
        )
    )

    if not normalized_insights:
        normalized_insights = (
            build_default_ai_insights(
                df
            )
        )

    normalized_recommendations = (
        normalize_excel_ai_items(
            recommendations
        )
    )

    if not normalized_recommendations:
        normalized_recommendations = (
            build_default_recommendations(
                df
            )
        )

    try:
        workbook = (
            create_professional_workbook()
        )

        # Sheet 1
        build_executive_summary_sheet(
            workbook=workbook,
            df=df,
            filename=safe_filename,
            metrics=calculated_metrics
        )

        # Sheet 2
        build_dataset_preview_sheet(
            workbook=workbook,
            df=df,
            filename=safe_filename,
            preview_rows=100
        )

        # Sheet 3
        build_complete_dataset_sheet(
            workbook=workbook,
            df=df,
            filename=safe_filename
        )

        # Sheet 4
        build_numeric_statistics_sheet(
            workbook=workbook,
            df=df,
            filename=safe_filename
        )

        # Sheet 5
        build_missing_value_sheet(
            workbook=workbook,
            df=df,
            filename=safe_filename
        )

        # Sheet 6
        build_ai_insights_sheet(
            workbook=workbook,
            df=df,
            filename=safe_filename,
            insights=normalized_insights
        )

        # Sheet 7
        build_recommendations_sheet(
            workbook=workbook,
            df=df,
            filename=safe_filename,
            recommendations=(
                normalized_recommendations
            )
        )

        if not workbook.worksheets:
            raise RuntimeError(
                (
                    "The professional workbook contains "
                    "no worksheets."
                )
            )

        workbook.active = 0

        workbook.properties.title = (
            "Professional Business "
            "Intelligence Report"
        )

        workbook.properties.subject = (
            f"Business intelligence analysis "
            f"for {safe_filename}"
        )

        workbook.properties.creator = (
            "AI Business Intelligence Platform"
        )

        workbook.properties.lastModifiedBy = (
            "AI Business Intelligence Platform"
        )

        workbook.properties.description = (
            (
                "Professional Excel workbook containing "
                "dataset metrics, data-quality analysis, "
                "descriptive statistics, AI insights and "
                "business recommendations."
            )
        )

        workbook.properties.keywords = (
            (
                "business intelligence, analytics, "
                "data quality, AI insights, recommendations"
            )
        )

        workbook.properties.category = (
            "Business Intelligence"
        )

        workbook.properties.created = (
            datetime.now()
        )

        workbook.properties.modified = (
            datetime.now()
        )

        workbook.save(
            output_path
        )

    except PermissionError as error:
        raise PermissionError(
            (
                "The Excel workbook could not be saved. "
                "Close the existing workbook if it is open "
                "in Microsoft Excel and try again."
            )
        ) from error

    except OSError as error:
        raise OSError(
            (
                "The professional Excel workbook could not "
                "be written to the selected output location."
            )
        ) from error

    except Exception as error:
        raise RuntimeError(
            (
                "Professional Excel report generation "
                f"failed: {str(error)}"
            )
        ) from error

    if not os.path.exists(
        output_path
    ):
        raise FileNotFoundError(
            (
                "Excel report generation completed without "
                "creating the expected output file."
            )
        )

    if os.path.getsize(
        output_path
    ) == 0:
        raise RuntimeError(
            (
                "The generated Excel workbook is empty."
            )
        )

    return output_path